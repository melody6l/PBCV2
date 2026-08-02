import os
import tempfile
import unittest

from document_registry import (
    all_available_files,
    document_change_detail,
    reconcile_source,
    register_source,
    resolve_export_link,
    resolve_preferred_path,
)


class DocumentRegistryTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = self.temp.name
        self.state = {}
        self.source = register_source(self.state, self.root)

    def tearDown(self):
        self.temp.cleanup()

    def write(self, relative, content):
        path = os.path.join(self.root, relative)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "wb") as stream:
            stream.write(content)
        return path

    def test_content_edit_creates_version_but_not_new_document(self):
        path = self.write("序时账.xlsx", b"first")
        first = reconcile_source(self.state, self.source, [path])
        document_id = next(iter(self.state["documents"]))
        self.assertEqual(first["new_files"], [path])

        self.write("序时账.xlsx", b"second")
        second = reconcile_source(self.state, self.source, [path])
        self.assertEqual(second["new_files"], [])
        self.assertEqual(second["updated_files"], [path])
        self.assertEqual(next(iter(self.state["documents"])), document_id)
        self.assertEqual(len(self.state["documents"][document_id]["versions"]), 2)

    def test_rename_keeps_identity_and_preferred_path(self):
        old_path = self.write("旧名.pdf", b"same")
        reconcile_source(self.state, self.source, [old_path])
        document_id = next(iter(self.state["documents"]))
        new_path = os.path.join(self.root, "新名.pdf")
        os.rename(old_path, new_path)

        changes = reconcile_source(self.state, self.source, [new_path])
        self.assertEqual(changes["moved_files"], [{"from": old_path, "to": new_path}])
        self.assertEqual(len(self.state["documents"]), 1)
        self.assertEqual(next(iter(self.state["documents"])), document_id)
        self.assertEqual(resolve_preferred_path(self.state, old_path), new_path)

    def test_same_content_at_two_paths_is_duplicate(self):
        first_path = self.write("公司1/序时账.xlsx", b"same")
        reconcile_source(self.state, self.source, [first_path])
        second_path = self.write("公司2/序时账.xlsx", b"same")

        changes = reconcile_source(self.state, self.source, [first_path, second_path])
        self.assertEqual(changes["duplicate_files"], [second_path])
        self.assertEqual(len(self.state["documents"]), 1)
        self.assertEqual(set(all_available_files(self.state)), {first_path, second_path})

    def test_identical_content_for_two_companies_is_not_merged(self):
        first_path = self.write("公司1/序时账.xlsx", b"same")
        second_path = self.write("公司2/序时账.xlsx", b"same")
        changes = reconcile_source(
            self.state,
            self.source,
            [first_path, second_path],
            company_names=["公司1", "公司2"],
        )
        self.assertEqual(len(self.state["documents"]), 2)
        self.assertEqual(changes["duplicate_files"], [])

    def test_cloud_export_uses_web_url_instead_of_cache_path(self):
        path = self.write("云端缓存/回函.pdf", b"pdf")
        cloud_source = register_source(
            self.state,
            self.root,
            source_type="microsoft_cloud",
            web_url="https://example.sharepoint.com/sites/a",
        )
        reconcile_source(
            self.state,
            cloud_source,
            [path],
            metadata_by_path={path: {"web_url": "https://example.sharepoint.com/file"}},
        )
        self.assertEqual(
            resolve_export_link(self.state, path),
            "https://example.sharepoint.com/file",
        )

    def test_excel_change_detail_reports_new_sheet_without_llm(self):
        from openpyxl import Workbook, load_workbook
        path = os.path.join(self.root, "序时账.xlsx")
        workbook = Workbook()
        workbook.active.title = "明细"
        workbook.save(path)
        reconcile_source(self.state, self.source, [path])

        workbook = load_workbook(path)
        workbook.create_sheet("调整汇总")
        workbook.save(path)
        reconcile_source(self.state, self.source, [path])
        detail = document_change_detail(self.state, path)
        self.assertTrue(detail["available"])
        self.assertIn("新增 Sheet：调整汇总", detail["changes"])


if __name__ == "__main__":
    unittest.main()
