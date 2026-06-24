"""Excel 读写处理 - 读取清单文件、导出匹配结果。"""

import json
import os
import uuid

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

META_HEADERS = {"row_uid", "source_key"}


def normalize_item_name(name):
    """归一化清单项名称，用于兼容旧版历史结果。"""
    return str(name).strip().replace("　", " ").lower()


def find_name_column(headers):
    """自动识别包含文件名称的列。"""
    name_keywords = ["文件名", "名称", "资料名称", "文件名称", "文档名称", "清单名称", "材料名称"]
    for i, header in enumerate(headers):
        if isinstance(header, str):
            for kw in name_keywords:
                if kw in header:
                    return i
    # 默认返回第二列，通常序号在第一列，名称在第二列
    visible_headers = [h for h in headers if h not in META_HEADERS]
    return 1 if len(visible_headers) > 1 else 0


def extract_path_from_link(cell):
    """从超链接单元格提取本地路径。"""
    target = cell.hyperlink.target if cell.hyperlink else cell.value
    if not target:
        return None
    target = str(target)
    if target.startswith("file:///"):
        return target[8:].replace("/", "\\")
    return None


def _column_index(headers, keyword):
    for i, header in enumerate(headers):
        if header and keyword in str(header):
            return i
    return None


def read_checklist(file_path):
    """读取 Excel 清单文件，返回列头和数据。"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = []
    hyperlink_rows = []
    for row in ws.iter_rows():
        rows.append([cell.value for cell in row])
        hyperlink_rows.append([extract_path_from_link(cell) for cell in row])

    if not rows:
        return {
            "headers": [],
            "data": [],
            "name_col_index": 0,
            "items": [],
            "history": {},
            "has_previous_results": False,
            "new_items": [],
            "existing_items": [],
            "prev_scanned_files": [],
            "prev_scanned_folders": [],
        }

    raw_headers = [str(h) if h else "" for h in rows[0]]
    generated_cols = {"核对结果", "文件超链接"}
    excluded_cols = {i for i, h in enumerate(raw_headers) if h in META_HEADERS or h in generated_cols}
    headers = [h for i, h in enumerate(raw_headers) if i not in excluded_cols]
    data = rows[1:]
    clean_data = [[cell for i, cell in enumerate(row) if i not in excluded_cols] for row in data]
    name_col_index = find_name_column(headers)

    status_col = _column_index(raw_headers, "核对结果")
    link_col = _column_index(raw_headers, "超链接")
    uid_col = raw_headers.index("row_uid") if "row_uid" in raw_headers else None
    source_key_col = raw_headers.index("source_key") if "source_key" in raw_headers else None
    prev_scanned_files = []
    prev_scanned_folders = []
    if "_pbc_meta" in wb.sheetnames:
        meta_ws = wb["_pbc_meta"]
        for row in meta_ws.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            if row[0] == "scanned_file":
                prev_scanned_files.append(str(row[1]))
            elif row[0] == "scanned_folder":
                prev_scanned_folders.append(str(row[1]))

    history = {}
    last_key = None
    for row_idx, row in enumerate(data):
        if not row:
            continue
        name = row[name_col_index] if len(row) > name_col_index else ""
        uid = row[uid_col] if uid_col is not None and len(row) > uid_col else ""
        source_key = row[source_key_col] if source_key_col is not None and len(row) > source_key_col else ""
        if name:
            status = row[status_col] if status_col is not None and len(row) > status_col else ""
            has_history_marker = bool(status or uid or source_key)
            if not has_history_marker:
                last_key = None
                continue
            source_key = str(source_key).strip() if source_key else normalize_item_name(name)
            key = str(uid).strip() if uid else source_key
            last_key = key
            if key not in history:
                normalized_status = str(status).strip() if status else "未匹配"
                if normalized_status == "未获取":
                    normalized_status = "未匹配"
                history[key] = {
                    "row_uid": str(uid).strip() if uid else "",
                    "source_key": source_key,
                    "status": normalized_status,
                    "matched_files": [],
                    "matched_names": [],
                    "matched_types": [],
                }
        elif last_key:
            key = last_key
        else:
            continue

        if link_col is not None and len(hyperlink_rows[row_idx + 1]) > link_col:
            matched_path = hyperlink_rows[row_idx + 1][link_col]
            if matched_path and matched_path not in history[key]["matched_files"]:
                history[key]["matched_files"].append(matched_path)
                history[key]["matched_names"].append(os.path.basename(matched_path))
                history[key]["matched_types"].append("文件夹" if os.path.isdir(matched_path) else "文件")

    items = []
    existing_items = []
    new_items = []
    for row_idx, row in enumerate(data):
        if row and len(row) > name_col_index:
            name = row[name_col_index]
            if name and str(name).strip():
                source_key = normalize_item_name(name)
                uid = row[uid_col] if uid_col is not None and len(row) > uid_col else ""
                item = {
                    "name": str(name).strip(),
                    "source_key": str(row[source_key_col]).strip() if source_key_col is not None and len(row) > source_key_col and row[source_key_col] else source_key,
                    "row_uid": str(uid).strip() if uid else "",
                    "row_index": row_idx + 1,
                }
                key = item["row_uid"] or item["source_key"]
                history_item = history.get(key) or history.get(item["source_key"])
                if history_item and history_item.get("status"):
                    item.update({
                        "status": history_item["status"],
                        "matched_files": history_item["matched_files"],
                        "matched_names": history_item["matched_names"],
                        "matched_types": history_item["matched_types"],
                        "need_match": history_item["status"] in ("未匹配", "部分获取", "待匹配"),
                    })
                    existing_items.append(item)
                else:
                    item.update({
                        "status": "待匹配" if history else "未匹配",
                        "matched_files": [],
                        "matched_names": [],
                        "matched_types": [],
                        "need_match": True,
                    })
                    new_items.append(item)
                items.append(item)

    wb.close()
    return {
        "headers": headers,
        "data": [[str(cell) if cell else "" for cell in row] for row in clean_data],
        "name_col_index": name_col_index,
        "items": items,
        "history": history,
        "has_previous_results": len(history) > 0,
        "new_items": new_items,
        "existing_items": existing_items,
        "prev_scanned_files": prev_scanned_files,
        "prev_scanned_folders": prev_scanned_folders,
    }


def export_results(results, headers, data, name_col_index, scan_root="", scanned_files=None, scanned_folders=None):
    """将匹配结果导出为 Excel 文件。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文件核对结果"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # 过滤掉所有数据行中都为空的列
    valid_cols = []
    for col_idx in range(len(headers)):
        header_val = headers[col_idx].strip() if col_idx < len(headers) else ""
        has_data = any(
            len(row) > col_idx and row[col_idx] and str(row[col_idx]).strip()
            for row in data
        )
        if header_val or has_data:
            valid_cols.append(col_idx)

    col_headers = [headers[c] for c in valid_cols] + ["核对结果", "文件超链接", "row_uid", "source_key"]
    for i, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 按清单名称建立索引，便于快速定位原始行
    checklist_lookup = {}
    for row in data:
        if len(row) > name_col_index:
            name_val = str(row[name_col_index]).strip() if row[name_col_index] else ""
            if name_val and name_val not in checklist_lookup:
                checklist_lookup[name_val] = row

    row_idx = 2
    for result in results:
        checklist_name = result.get("checklist_name", "")
        orig_row = checklist_lookup.get(checklist_name, [])
        matched_files = result.get("matched_files", []) or []
        matched_names = result.get("matched_names", []) or []
        n_rows = max(1, len(matched_files))
        row_uid = result.get("row_uid") or str(uuid.uuid4())
        source_key = result.get("source_key") or normalize_item_name(checklist_name)

        start_row = row_idx
        end_row = row_idx + n_rows - 1

        for sub_idx in range(n_rows):
            current_row = row_idx + sub_idx

            # 原始列只在第一行写入，后续通过纵向合并展示
            if sub_idx == 0 and orig_row:
                for col_num, col_idx in enumerate(valid_cols, 1):
                    cell_val = orig_row[col_idx] if len(orig_row) > col_idx else ""
                    cell = ws.cell(row=current_row, column=col_num, value=cell_val)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 核对结果列只在第一行写入，后续通过纵向合并展示
            if sub_idx == 0:
                status_col = len(valid_cols) + 1
                status_cell = ws.cell(row=current_row, column=status_col, value=result["status"])
                status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if result["status"] == "已获取":
                    status_cell.fill = green_fill
                elif result["status"] == "部分获取":
                    status_cell.fill = partial_fill
                else:
                    status_cell.fill = red_fill

            # 文件超链接列：每个文件独占一行
            link_col = len(valid_cols) + 2
            if sub_idx < len(matched_files):
                file_name = matched_names[sub_idx] if sub_idx < len(matched_names) else ""
                file_path = matched_files[sub_idx]
                link_url = "file:///" + file_path.replace("\\", "/")
                link_cell = ws.cell(row=current_row, column=link_col, value=file_name)
                link_cell.hyperlink = link_url
                link_cell.font = Font(color="0563C1", underline="single")
                link_cell.alignment = Alignment(vertical="center")

            uid_col = len(valid_cols) + 3
            key_col = len(valid_cols) + 4
            ws.cell(row=current_row, column=uid_col, value=row_uid if sub_idx == 0 else "")
            ws.cell(row=current_row, column=key_col, value=source_key if sub_idx == 0 else "")

        if n_rows > 1:
            # 原始列纵向合并
            for col_num in range(1, len(valid_cols) + 1):
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_num,
                    end_row=end_row,
                    end_column=col_num,
                )
                merged_cell = ws.cell(row=start_row, column=col_num)
                merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 状态列纵向合并
            status_col = len(valid_cols) + 1
            ws.merge_cells(
                start_row=start_row,
                start_column=status_col,
                end_row=end_row,
                end_column=status_col,
            )
            status_cell = ws.cell(row=start_row, column=status_col)
            status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_idx += n_rows

    ws.column_dimensions[get_column_letter(len(valid_cols) + 3)].hidden = True
    ws.column_dimensions[get_column_letter(len(valid_cols) + 4)].hidden = True

    meta_ws = wb.create_sheet("_pbc_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["scan_root", scan_root])
    meta_ws.append(["result_count", len(results)])
    for path in scanned_files or []:
        meta_ws.append(["scanned_file", path])
    for path in scanned_folders or []:
        meta_ws.append(["scanned_folder", path])
    for result in results:
        meta_ws.append([
            result.get("row_uid", ""),
            result.get("source_key", ""),
            result.get("status", ""),
            json.dumps(result.get("matched_files", []), ensure_ascii=False),
        ])

    # 设置列宽
    for i in range(1, len(col_headers) + 1):
        if i <= len(valid_cols):
            ws.column_dimensions[get_column_letter(i)].width = 18
        elif i == len(valid_cols) + 1:
            ws.column_dimensions[get_column_letter(i)].width = 12
        else:
            ws.column_dimensions[get_column_letter(i)].width = 36

    os.makedirs("exports", exist_ok=True)
    output_path = "exports/result.xlsx"
    wb.save(output_path)
    wb.close()
    return output_path


def build_browse_items(scanned_files, scanned_folders, scan_root, match_results):
    """构建带动态文件夹层级和关联需求的资料列表。"""
    path_requirements = {}
    for result in match_results or []:
        requirement = {
            "index": result.get("index"),
            "pbc_name": result.get("pbc_name", ""),
        }
        for path in result.get("matched_files", []) or []:
            requirements = path_requirements.setdefault(path, [])
            if requirement not in requirements:
                requirements.append(requirement)

    paths = list(dict.fromkeys((scanned_folders or []) + (scanned_files or [])))
    if scan_root:
        paths.sort(key=lambda path: os.path.relpath(path, scan_root).lower())
    else:
        paths.sort(key=str.lower)

    items = []
    folder_levels = 0
    for path in paths:
        relative_path = os.path.relpath(path, scan_root) if scan_root else os.path.basename(path)
        parts = [part for part in relative_path.split(os.sep) if part not in ("", ".")]
        if not parts or parts[0] == "..":
            parts = [os.path.basename(path)]
        folder_parts = parts[:-1]
        matched_requirements = path_requirements.get(path, [])
        folder_levels = max(folder_levels, len(folder_parts))
        is_dir = os.path.isdir(path)
        if is_dir:
            continue
        items.append({
            "seq": len(items) + 1,
            "folder_parts": folder_parts,
            "filename": parts[-1],
            "path": path,
            "is_dir": False,
            "is_matched": bool(matched_requirements),
            "matched_requirements": matched_requirements,
        })
    return items, folder_levels

def export_checklist_two_sheets(items, company_names, match_results, file_renames=None, scanned_files=None, scanned_folders=None, scan_root=""):
    """导出包含矩阵视图、清单视图和资料浏览视图的Excel文件。

    Args:
        items: 清单项列表（每个PBC条目独立，含 company_status）
        company_names: 公司简称列表
        match_results: 匹配结果列表（含 index, status, matched_files, company_coverage）
        file_renames: 文件重命名映射 {原始路径: 新名称}

    Returns:
        output_path: 生成的Excel文件路径
    """
    import os
    if file_renames is None:
        file_renames = {}
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ========== Sheet 1: 矩阵视图 ==========
    ws_matrix = wb.active
    ws_matrix.title = "矩阵视图"

    matrix_headers = ["序号", "科目", "所需PBC", "需求资料"] + company_names
    for col_idx, h in enumerate(matrix_headers, 1):
        cell = ws_matrix.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        ws_matrix.cell(row=row_idx, column=1, value=item.get("seq", row_idx - 1)).alignment = center_align
        ws_matrix.cell(row=row_idx, column=1).border = thin_border
        ws_matrix.cell(row=row_idx, column=2, value=item.get("subject", "")).alignment = center_align
        ws_matrix.cell(row=row_idx, column=2).border = thin_border
        ws_matrix.cell(row=row_idx, column=3, value=item.get("pbc_name", "")).alignment = center_align
        ws_matrix.cell(row=row_idx, column=3).border = thin_border
        ws_matrix.cell(row=row_idx, column=4, value=item.get("demand_name", "")).alignment = center_align
        ws_matrix.cell(row=row_idx, column=4).border = thin_border

        company_status = item.get("company_status", {})
        for ci, cName in enumerate(company_names, 5):
            status = company_status.get(cName, "N")
            cell = ws_matrix.cell(row=row_idx, column=ci, value=status)
            cell.alignment = center_align
            cell.border = thin_border
            if status == "Y":
                cell.fill = green_fill
            elif status == "不完整":
                cell.fill = orange_fill
            elif status == "N/A":
                cell.fill = gray_fill
            else:
                cell.fill = red_fill

    ws_matrix.column_dimensions["A"].width = 8
    ws_matrix.column_dimensions["B"].width = 16
    ws_matrix.column_dimensions["C"].width = 30
    ws_matrix.column_dimensions["D"].width = 30
    for ci in range(5, 5 + len(company_names)):
        ws_matrix.column_dimensions[get_column_letter(ci)].width = 10

    # ========== Sheet 2: 清单视图 ==========
    ws_list = wb.create_sheet("清单视图")

    list_headers = ["科目", "所需PBC", "需求资料", "公司", "是否获取", "文件"]
    for col_idx, h in enumerate(list_headers, 1):
        cell = ws_list.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 建立 match_results 索引
    match_lookup = {}
    if match_results:
        for r in match_results:
            match_lookup[r["index"]] = r

    row_idx = 2
    for item in items:
        for cName in company_names:
            status = item.get("company_status", {}).get(cName, "N")
            if status == "不完整":
                status_text = "不完整"
            elif status == "Y":
                status_text = "Y"
            elif status == "N/A":
                status_text = "N/A"
            else:
                status_text = "N"

            # 辅助函数：写入一行上下文列（科目/PBC/需求资料/公司/状态）
            def write_context_columns(r):
                ws_list.cell(row=r, column=1, value=item.get("subject", "")).alignment = center_align
                ws_list.cell(row=r, column=1).border = thin_border
                ws_list.cell(row=r, column=2, value=item.get("pbc_name", "")).alignment = center_align
                ws_list.cell(row=r, column=2).border = thin_border
                ws_list.cell(row=r, column=3, value=item.get("demand_name", "")).alignment = center_align
                ws_list.cell(row=r, column=3).border = thin_border
                ws_list.cell(row=r, column=4, value=cName).alignment = center_align
                ws_list.cell(row=r, column=4).border = thin_border
                sc = ws_list.cell(row=r, column=5, value=status_text)
                sc.alignment = center_align
                sc.font = Font(bold=True)
                sc.border = thin_border
                if status == "Y":
                    sc.fill = green_fill
                elif status == "不完整":
                    sc.fill = orange_fill
                else:
                    sc.fill = red_fill

            # 收集文件路径
            file_paths = []
            if status == "Y":
                ri = item.get("row_index")
                mr = match_lookup.get(ri)
                seen_paths = set()
                if mr:
                    company_info = mr.get("company_coverage", {}).get(cName, {})
                    if company_info:
                        fp_list = (company_info.get("files", []) or []) + (company_info.get("folders", []) or [])
                    else:
                        fp_list = mr.get("matched_files", []) or []
                    for fp in fp_list:
                        if fp not in seen_paths:
                            seen_paths.add(fp)
                            file_paths.append(fp)

            if not file_paths:
                # 无文件：写一行空文件列
                write_context_columns(row_idx)
                file_cell = ws_list.cell(row=row_idx, column=6, value="")
                file_cell.alignment = Alignment(wrap_text=True, vertical="top")
                file_cell.border = thin_border
                row_idx += 1
            else:
                # 多文件：每个文件独占一行
                for fi, fp in enumerate(file_paths):
                    fname = file_renames.get(fp, os.path.basename(fp))
                    link_url = "file:///" + fp.replace("\\", "/")
                    write_context_columns(row_idx)
                    link_cell = ws_list.cell(row=row_idx, column=6, value=fname)
                    link_cell.hyperlink = link_url
                    link_cell.font = Font(color="0563C1", underline="single")
                    link_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    link_cell.border = thin_border
                    row_idx += 1

    ws_list.column_dimensions["A"].width = 16
    ws_list.column_dimensions["B"].width = 30
    ws_list.column_dimensions["C"].width = 30
    ws_list.column_dimensions["D"].width = 12
    ws_list.column_dimensions["E"].width = 10
    ws_list.column_dimensions["F"].width = 50

    # ========== Sheet 3: 资料浏览视图 ==========
    browse_items, folder_levels = build_browse_items(
        scanned_files, scanned_folders, scan_root, match_results
    )
    ws_browse = wb.create_sheet("资料浏览视图")
    level_names = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    browse_headers = [
        f"{level_names[level - 1] if level <= len(level_names) else level}级文件夹"
        for level in range(1, folder_levels + 1)
    ]
    browse_headers += ["文件名", "匹配状态", "关联需求"]

    for col_idx, header in enumerate(browse_headers, 1):
        cell = ws_browse.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, browse_item in enumerate(browse_items, 2):
        row_fill = green_fill if browse_item["is_matched"] else red_fill
        folder_parts = browse_item["folder_parts"]
        requirement_text = "、".join(
            requirement["pbc_name"]
            for requirement in browse_item["matched_requirements"]
        )
        values = folder_parts + [""] * (folder_levels - len(folder_parts))
        values += [
            browse_item["filename"],
            "已匹配" if browse_item["is_matched"] else "未匹配",
            requirement_text,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws_browse.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        filename_col = folder_levels + 1
        filename_cell = ws_browse.cell(row=row_idx, column=filename_col)
        filename_cell.hyperlink = "file:///" + browse_item["path"].replace("\\", "/")
        filename_cell.font = Font(color="0563C1", underline="single")

    for col_idx in range(1, folder_levels + 1):
        ws_browse.column_dimensions[get_column_letter(col_idx)].width = 20
    ws_browse.column_dimensions[get_column_letter(folder_levels + 1)].width = 36
    ws_browse.column_dimensions[get_column_letter(folder_levels + 2)].width = 14
    ws_browse.column_dimensions[get_column_letter(folder_levels + 3)].width = 50
    ws_browse.freeze_panes = "A2"

    os.makedirs("exports", exist_ok=True)
    output_path = "exports/PBC需求清单_三视图.xlsx"
    wb.save(output_path)
    wb.close()
    return output_path
