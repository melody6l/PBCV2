"""PBC需求清单模板处理 - 读取模板、生成清单、解析用户填写结果"""

import os
import uuid
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from excel_handler import normalize_item_name


TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "资料表.xlsx")


def read_template():
    """读取资料表模板，返回科目列表和资料项列表"""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["资料表"]

    subjects = set()
    items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        # 资料表.xlsx 实际结构: A列=序号, B列=科目, C列=所需PBC
        if not row or len(row) < 2 or not row[1]:
            continue
        seq = int(row[0]) if row[0] is not None else row_idx
        subject = str(row[1]).strip() if row[1] else ""
        pbc_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not subject:
            continue
        subjects.add(subject)
        items.append({
            "row_index": row_idx,
            "seq": seq,
            "subject": subject,
            "pbc_name": pbc_name,
        })

    wb.close()
    return {
        "subjects": sorted(subjects),
        "items": items,
    }


def generate_checklist(selected_subjects, company_full_name="", company_short_name=""):
    """
    根据用户选择的科目生成PBC需求清单Excel文件，返回文件路径。

    selected_subjects: ["All", "货币资金", "固定资产", ...]
    """
    wb_template = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_template = wb_template["资料表"]

    # 读取模板数据（资料表.xlsx 实际结构: A列=序号, B列=科目, C列=所需PBC）
    all_items = []
    for row in ws_template.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue
        subject = str(row[1]).strip() if row[1] else ""
        pbc_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not subject:
            continue
        if subject in selected_subjects:
            all_items.append({
                "subject": subject,
                "pbc_name": pbc_name,
            })

    wb_template.close()

    # 创建输出工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "核对工作台"

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头: 序号 | 科目 | 所需PBC | 需求资料
    headers = ["序号", "科目", "所需PBC", "需求资料"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 填入数据行
    for row_idx, item in enumerate(all_items, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=2, value=item["subject"]).alignment = center_align
        ws.cell(row=row_idx, column=2).border = thin_border
        ws.cell(row=row_idx, column=3, value=item["pbc_name"]).alignment = center_align
        ws.cell(row=row_idx, column=3).border = thin_border
        # D列 默认=C列内容
        d_cell = ws.cell(row=row_idx, column=4, value=item["pbc_name"])
        d_cell.alignment = center_align
        d_cell.border = thin_border

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 42

    # 公司名称 sheet
    ws_company = wb.create_sheet("公司名称")
    ws_company.cell(row=1, column=1, value="公司全称").fill = header_fill
    ws_company.cell(row=1, column=1).font = header_font
    ws_company.cell(row=1, column=1).alignment = center_align
    ws_company.cell(row=1, column=2, value="简称").fill = header_fill
    ws_company.cell(row=1, column=2).font = header_font
    ws_company.cell(row=1, column=2).alignment = center_align
    if company_full_name:
        ws_company.cell(row=2, column=1, value=company_full_name).alignment = center_align
    if company_short_name:
        ws_company.cell(row=2, column=2, value=company_short_name).alignment = center_align
    ws_company.column_dimensions["A"].width = 30
    ws_company.column_dimensions["B"].width = 14

    os.makedirs("uploads", exist_ok=True)
    output_path = os.path.join("uploads", "PBC需求清单_待填写.xlsx")
    wb.save(output_path)
    wb.close()
    return output_path


def read_user_checklist(file_path):
    """
    读取用户填写后的PBC需求清单，解析为前端预览数据。

    返回:
    {
        "companies": [{"full_name": "...", "short_name": "..."}],
        "items": [
            {
                "seq": 1,
                "subject": "All",
                "pbc_name": "科目余额表",
                "demand_name": "PRC Report",
                "company_status": {"CN02": "Y", "CN03": "N", ...},
            }
        ],
        "headers": ["序号", "科目", "所需PBC", "需求资料", "CN02", "CN03", ...]
    }
    """
    wb = openpyxl.load_workbook(file_path)

    # 读取公司名称
    companies = []
    if "公司名称" in wb.sheetnames:
        ws_company = wb["公司名称"]
        for row in ws_company.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                companies.append({
                    "full_name": str(row[0]).strip(),
                    "short_name": str(row[1]).strip() if row[1] else "",
                })

    # 读取核对工作台
    ws = None
    for sname in wb.sheetnames:
        if sname in ("资料预览", "核对工作台", "示例"):
            ws = wb[sname]
            break
    if ws is None:
        ws = wb.active

    headers = []
    items = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
        raw_vals = [c.value for c in row]
        if row_idx == 0:
            headers = [str(v) if v else "" for v in raw_vals]
            continue

        # A=序号, B=科目, C=所需PBC, D=需求资料, E~=各公司状态
        if not raw_vals[1] and not raw_vals[2]:
            continue

        # 检查是否有 _deleted 标记（最后一个非None值为 True 布尔值）
        has_deleted = any(
            isinstance(v, bool) and v is True
            for v in raw_vals[4:] if v is not None
        )
        if has_deleted:
            continue

        company_status = {}
        for ci in range(4, len(raw_vals)):
            val = raw_vals[ci]
            col_header = headers[ci] if ci < len(headers) else ""
            if col_header and val is not None:
                sv = str(val).strip().upper()
                if sv in ("Y", "YES", "1", "TRUE"):
                    company_status[col_header] = "Y"
                elif sv in ("N/A", "NA"):
                    company_status[col_header] = "N/A"
                elif sv in ("不完整", "INCOMPLETE", "I"):
                    company_status[col_header] = "不完整"
                elif sv in ("N", "NO", "0", "FALSE"):
                    company_status[col_header] = "N"
                elif sv and str(val).strip() == "不完整":
                    company_status[col_header] = "不完整"
                elif sv:
                    company_status[col_header] = "N"

        items.append({
            "row_index": row_idx,
            "seq": int(raw_vals[0]) if raw_vals[0] is not None else row_idx,
            "subject": str(raw_vals[1]).strip() if len(raw_vals) > 1 and raw_vals[1] else "",
            "pbc_name": str(raw_vals[2]).strip() if len(raw_vals) > 2 and raw_vals[2] else "",
            "demand_name": str(raw_vals[3]).strip() if len(raw_vals) > 3 and raw_vals[3] else "",
            "company_status": company_status,
        })

    wb.close()

    company_names = [c["short_name"] or c["full_name"] for c in companies]
    return {
        "companies": companies,
        "company_names": company_names,
        "items": items,
        "headers": headers[:4] + company_names,
    }


def update_cell_status(file_path, row_index, company_name, status):
    """更新某个单元格的获取状态（Y/N/N/A）"""
    wb = openpyxl.load_workbook(file_path)
    ws = None
    for sname in wb.sheetnames:
        if sname in ("资料预览", "核对工作台", "示例"):
            ws = wb[sname]
            break
    if ws is None:
        ws = wb.active

    # 找到公司名称所在列
    target_col = None
    for col_idx in range(5, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val and str(val).strip() == company_name:
            target_col = col_idx
            break

    if target_col is None:
        wb.close()
        raise ValueError(f"未找到公司列: {company_name}")

    cell = ws.cell(row=row_index, column=target_col)

    # 颜色样式
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    cell.value = status
    cell.alignment = center_align
    if status == "Y":
        cell.fill = green_fill
    elif status == "不完整":
        cell.fill = orange_fill
    elif status == "N/A":
        cell.fill = gray_fill
    else:
        cell.fill = red_fill

    wb.save(file_path)
    wb.close()
    return True


def export_checklist_with_status(file_path):
    """导出当前填写好状态的Excel文件"""
    return file_path


def generate_checklist_from_memory(tpl_data):
    """
    从内存中的 checklist_template 数据生成 Excel 文件，供下载用。

    tpl_data 结构（与 read_user_checklist 返回值一致）:
    {
        "companies": [{"full_name": "...", "short_name": "..."}],
        "company_names": ["CN02", "CN03", ...],
        "items": [
            {
                "row_index": 2,
                "seq": 1,
                "subject": "货币资金",
                "pbc_name": "科目余额表",
                "demand_name": "PRC Report",
                "company_status": {"CN02": "Y", "CN03": "N", ...},
            }
        ],
        "headers": [...]
    }
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "核对工作台"

    # 样式定义（与 generate_checklist 保持一致）
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    companies = tpl_data.get("companies", [])
    company_names = tpl_data.get("company_names", [])
    items = tpl_data.get("items", [])

    # 表头: 序号 | 科目 | 所需PBC | 需求资料 | 公司1 | 公司2 | ...
    base_headers = ["序号", "科目", "所需PBC", "需求资料"]
    headers = base_headers + company_names
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    for row_idx, item in enumerate(items, 2):
        seq = item.get("seq", row_idx - 1)
        subject = item.get("subject", "")
        pbc_name = item.get("pbc_name", "")
        demand_name = item.get("demand_name", "") or pbc_name
        company_status = item.get("company_status", {})

        vals = [seq, subject, pbc_name, demand_name]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border

        # 公司状态列（从第5列开始）
        for ci, cn in enumerate(company_names):
            cell = ws.cell(row=row_idx, column=5 + ci, value=company_status.get(cn, ""))
            cell.alignment = center_align
            cell.border = thin_border
            status = company_status.get(cn, "")
            if status == "Y":
                cell.fill = green_fill
            elif status == "不完整":
                cell.fill = orange_fill
            elif status == "N/A":
                cell.fill = gray_fill
            elif status == "N":
                cell.fill = red_fill

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 42
    for ci in range(len(company_names)):
        col_letter = get_column_letter(5 + ci)
        ws.column_dimensions[col_letter].width = 14

    # 公司名称 sheet
    ws_company = wb.create_sheet("公司名称")
    ws_company.cell(row=1, column=1, value="公司全称").fill = header_fill
    ws_company.cell(row=1, column=1).font = header_font
    ws_company.cell(row=1, column=1).alignment = center_align
    ws_company.cell(row=1, column=2, value="简称").fill = header_fill
    ws_company.cell(row=1, column=2).font = header_font
    ws_company.cell(row=1, column=2).alignment = center_align
    for ri, company in enumerate(companies, 2):
        ws_company.cell(row=ri, column=1, value=company.get("full_name", "")).alignment = center_align
        ws_company.cell(row=ri, column=2, value=company.get("short_name", "")).alignment = center_align
    ws_company.column_dimensions["A"].width = 30
    ws_company.column_dimensions["B"].width = 14

    # 保存
    os.makedirs("uploads", exist_ok=True)
    output_path = os.path.join("uploads", "PBC需求清单_待填写.xlsx")
    wb.save(output_path)
    wb.close()
    return output_path


# ====== 行管理：新增 / 编辑 / 删除 ======

def _find_checklist_sheet(wb):
    """查找核对清单工作表，兼容多种 sheet 名称"""
    for name in ["核对工作台", "资料预览", "示例"]:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def add_row_to_checklist(file_path, subject, pbc_name, demand_name,
                         position=None, ref_row_index=None):
    """
    向核对清单 Excel 中追加/插入一行。
    返回新行的 worksheet 行号（即 row_index）。
    """
    wb = openpyxl.load_workbook(file_path)
    ws = _find_checklist_sheet(wb)

    # 确定插入位置
    if position == "top" and ref_row_index:
        insert_row = ref_row_index
        ws.insert_rows(insert_row)
    elif position == "bottom" and ref_row_index:
        insert_row = ref_row_index + 1
        ws.insert_rows(insert_row)
    else:
        insert_row = ws.max_row + 1

    company_start_col = 5  # E列开始是公司状态
    company_count = ws.max_column - 4  # 公司列数

    # 样式定义（与 generate_checklist 一致）
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    seq_cell = ws.cell(row=insert_row, column=1, value=insert_row - 1)
    seq_cell.alignment = center_align
    seq_cell.border = thin_border

    subject_cell = ws.cell(row=insert_row, column=2, value=subject)
    subject_cell.alignment = center_align
    subject_cell.border = thin_border

    pbc_cell = ws.cell(row=insert_row, column=3, value=pbc_name)
    pbc_cell.alignment = center_align
    pbc_cell.border = thin_border

    demand_cell = ws.cell(row=insert_row, column=4, value=demand_name)
    demand_cell.alignment = center_align
    demand_cell.border = thin_border

    # 初始化公司状态为 "N"
    for col_idx in range(company_start_col, company_start_col + company_count):
        cell = ws.cell(row=insert_row, column=col_idx, value="N")
        cell.alignment = center_align
        cell.border = thin_border

    wb.save(file_path)
    wb.close()
    return insert_row


# 字段名 → Excel 列号映射
FIELD_COL_MAP = {"subject": 2, "pbc_name": 3, "demand_name": 4}


def edit_row_in_checklist(file_path, row_index, field, value):
    """编辑核对清单 Excel 中指定行的某个字段"""
    col = FIELD_COL_MAP.get(field)
    if col is None:
        raise ValueError(f"未知字段: {field}")

    wb = openpyxl.load_workbook(file_path)
    ws = _find_checklist_sheet(wb)
    ws.cell(row=row_index, column=col, value=value)
    wb.save(file_path)
    wb.close()


def delete_row_from_checklist(file_path, row_index):
    """
    逻辑删除核对清单中的一行。
    在最后一列写入 _deleted 标记，导出时过滤。
    避免物理删除导致 row_index 偏移影响匹配结果。
    """
    wb = openpyxl.load_workbook(file_path)
    ws = _find_checklist_sheet(wb)
    # 在最后一列+1 写入删除标记
    marker_col = ws.max_column + 1
    ws.cell(row=row_index, column=marker_col, value=True)
    wb.save(file_path)
    wb.close()
